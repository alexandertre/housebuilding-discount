# Import openyxl module
import openpyxl
import json
import xlrd
from itertools import groupby

f = open('rates.txt', 'w', encoding='utf-8')

rates = './data/rates.xlsx'
costs = './data/costs.xlsx'
materials = './data/materials.xlsx'
# Define variable to load the wookbook
costs_list = []


def readxlsx(data):
    global costs_list

    if (data):
        wookbook = openpyxl.load_workbook(data)
        # Define variable to read the active sheet:
        worksheet = wookbook.active
        # Iterate the loop to read the cell values

        # for col in worksheet.iter_cols(1, worksheet.max_column):

        for col in worksheet.iter_rows(min_row=2):
            rate = {}
            rate['rate_id'] = int(col[0].value)
            string = str(col[1].value)
            string = string.replace('"', '')
            string = string.replace("'", "")
            rate['name'] = string
            costs_list.append(rate)
def test(data):
    global costs_list
    if (data):
        wookbook = openpyxl.load_workbook(data)
        # Define variable to read the active sheet:
        worksheet = wookbook.active
        # Iterate the loop to read the cell values
        print('rows', worksheet.max_row)
        print('cols', worksheet.max_column)
        for col in worksheet.iter_rows(1, worksheet.max_row):
            print(col[0].value)


readxlsx(rates)
result = str(costs_list)
final = result.replace("'", '"')
f.write(final)