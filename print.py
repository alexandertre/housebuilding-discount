# Import openyxl module
import openpyxl
import json
import xlrd
from itertools import groupby

f = open('results.txt', 'w', encoding='utf-8')

rates = './data/rates.xlsx'
costs = './data/costs.xlsx'
materials = './data/materials.xlsx'
# Define variable to load the wookbook
costs_list = []
#new_costs_list1 = str(costs_list)

newcosts = {'costs': costs_list}


def readxlsx(data):
    global costs_list

    if (data):
        wookbook = openpyxl.load_workbook(data)
        # Define variable to read the active sheet:
        worksheet = wookbook.active
        # Iterate the loop to read the cell values

        # for col in worksheet.iter_cols(1, worksheet.max_column):

        for col in worksheet.iter_rows(min_row=2):
            new_materials = []
            if col[14].value:
                new_materials.append({'material_id': col[14].value,
                'price': str(col[13].value)})
                #material = '[{"material_id":'+ str(col[14].value) + ', "price": ' + str(col[13].value) + '}]'
                #new_material = material.replace('"[', '[')
                #new2_material = new_material.replace(']"', ']')
                #material=new2_material
            else:
                new_materials.append("null")
                #material = material.replace('"', '')

            rates = []
            detail = {}
            rate = {'rate_id': col[9].value, 'price': float(col[11].value), 'materials': new_materials}
            rates.append(rate)
            detail['rates'] = rates
            costs_list.append({'building_category_id': col[3].value,
                               'technology_type_id': col[4].value,
                               'koef_etazh': col[5].value,
                               'koef_m2_doma': col[6].value,
                               'koef_m2_uchastka': col[7].value,
                               'location_id': col[1].value,
                               'price': float(col[8].value),
                               'detail': detail
                               })
            # print(col[i].value, end="\t\t")
            #print('success')

    else:
        print('reading file error')


def test(data):
    global costs_list
    if (data):
        wookbook = openpyxl.load_workbook(data)
        # Define variable to read the active sheet:
        worksheet = wookbook.active
        # Iterate the loop to read the cell values
        print('rows', worksheet.max_row)
        print('cols', worksheet.max_column)
        for col in worksheet.iter_rows(1, worksheet.max_row):
            print(col[0].value)


readxlsx(costs)
result = str(newcosts)
result = result.replace('"[', '[')
result = result.replace(']"', ']')
result = result.replace('"null"', 'null')
result = result.replace("'null'", "null")
result = result.replace("'", '"')

material_list_result = []
materials_list=[]
def read_obj(data):
    global materials_list
    new_obj = {}
    for obj in data:
        detail = obj['detail']
        cat_price = obj['price']
        materials = detail['rates'][0]
        material_obj = {}
        material_obj['cat_price'] = cat_price
        single_material_array = []
        single_material = materials['materials'][0]
        one_material = {}
        one_material_obj = {}
        one_material_obj['material_id'] = single_material['material_id']
        one_material_obj['price'] = single_material['price']
        single_material_array.append(one_material_obj)
        material_obj['single_material'] = single_material_array
        materials_list.append(material_obj)

pre_result_rate_material_list = []
def resultmaterials(data):
    global pre_result_rate_material_list
    print(data)
    new_obj = {}
    i = 0
    for element in data[:20]:
        i += 1
        #print('element from array 1'+str(element))
        #print(str(i)+' cat_price: '+str(element['cat_price']))
        if(len(pre_result_rate_material_list) > 0):
            #print('select from array 2')
            for f_element in pre_result_rate_material_list:
                #f_element['single_rate']['materials'] = [el for el, _ in groupby(f_element['single_rate']['materials'])]
                #print('element from array 2' + str(f_element))
                if(int(f_element['single_rate']['rate_id']) == int(element['single_rate']['rate_id']) and float(f_element['cat_price']) == float(element['cat_price'])):
                    #print('matched')
                    #new_materials_array = [el for el, _ in groupby(f_element['single_rate']['materials'])]
                    for material in f_element['single_rate']['materials']:
                        #print('material from array 2' + str(material))
                        if(material['material_id'] != element['single_rate']['materials'][0]['material_id']):
                            f_element['single_rate']['materials'].append(element['single_rate']['materials'][0].copy())
                            #print('added new material to array 2 with' + str(f_element['single_rate']['materials']) + ' new one' +  str(material))
                            #print('updated' + str(element['single_rate']['materials'][0]))
                        else:
                            print('missed')
                else:
                    obj = {}
                    obj['cat_price'] = float(element['cat_price'])
                    obj['single_rate'] = element['single_rate'].copy()
                    pre_result_rate_material_list.append(obj)
                    obj = ''
                    #print(str(i) + ' iteration & added new one' + str(obj))
        else:
            new_obj['cat_price'] = float(element['cat_price'])
            new_obj['single_rate'] = element['single_rate'].copy()
            pre_result_rate_material_list.append(new_obj)
            #print('added first element')

pre_result_list_rates = []
def rateslist(data):
    global pre_result_list_rates
    for rate in data:
        new_obj = {}
        new_obj['cat_price'] = rate['price']
        new_obj['single_rate'] = rate['detail']['rates'][0]
        pre_result_list_rates.append(new_obj)
it = 0
third_array_loop_value = ''
second_array_loop_value = ''
duplicate =''

rate_array = []
def second_array_loop(element):
    global pre_result_rate_material_list, second_array_loop_value, third_array_loop_value
    i = 0

    def third_array_loop(element_loop1, element_loop2):
        global pre_result_rate_material_list, it, third_array_loop_value, duplicate, rate_array
        it += 1
        # print(str(it) + ' it')
        if (float(element_loop1['cat_price']) == float(element_loop2['cat_price']) and int(
                element_loop1['single_rate']['rate_id']) == int(element_loop2['single_rate']['rate_id'])):
            # print('got one')
            rate_array.append('duplicate')
            new_array = []
            for material in element_loop2['single_rate']['materials']:
                if (material['material_id'] == element_loop1['single_rate']['materials'][0]['material_id']):
                    marker = 'already'
                    # print('already')
                    new_array.append('already')
                    third_array_loop_value = marker
                else:
                    marker = 'new'
                    # print('new')
                    new_array.append('new')
                    third_array_loop_value = marker
            for new in new_array:
                if (new == 'new'):
                    material['material_id'] == element_loop1['single_rate']['materials'][0]['material_id']
        else:
            marker = 'false'
            rate_array.append('new')
            duplicate = marker
    for second_array_element in pre_result_rate_material_list:
       i += 1
       #print(i)
       #print(str(len(pre_result_rate_material_list)) + ' elements')

       check = third_array_loop(element, second_array_element)
       #if(third_array_loop_value == 'new'):
           #marker = 'new'
           #second_array_element['single_rate']['materials'].append(element['single_rate']['materials'][0])
       if(third_array_loop_value == 'false'):
           marker = 'false'
           second_array_loop_value = marker

    for new_string in rate_array:
        if(new_string == 'new'):
            pre_result_rate_material_list.append(new_string)

def result_rate_materials(data):
    global pre_result_rate_material_list
    for element in data[:10]:
        if(len(pre_result_rate_material_list) > 0):
            #print('second array')
            check = second_array_loop(element)
            if (duplicate == 'false'):
                print('added new')
        else:
            pre_result_rate_material_list.append(element)

def new_result_rate_materials(data):
    global pre_result_rate_material_list
    for element in data:
        if(len(pre_result_rate_material_list) > 0):
            first_array = []
            for second_array_element in pre_result_rate_material_list:
                if (float(element['cat_price']) == float(second_array_element['cat_price']) and int(element['single_rate']['rate_id']) == int(second_array_element['single_rate']['rate_id'])):
                    first_array.append('duplicate')
                    material_array = []
                    for material in second_array_element['single_rate']['materials']:
                        if (material['material_id'] == element['single_rate']['materials'][0]['material_id']):
                            material_array.append('duplicate')
                    if(len(material_array) < 1):
                        element['single_rate']['materials'][0]['price'] = float(element['single_rate']['materials'][0]['price'])
                        second_array_element['single_rate']['materials'].append(element['single_rate']['materials'][0])

                else:
                    first_array.append('new')
            control_array = []
            for first_array_element in first_array:
                if(first_array_element == 'duplicate'):
                    control_array.append('duplicate')
            if(len(control_array) < 1):
                pre_result_rate_material_list.append(element)
        else:
            pre_result_rate_material_list.append(element)

pre_result_rates = []
def mulirates(data):
    global pre_result_rates
    for element in data:
        if(len(pre_result_rates) > 0):
            first_array = []
            for second_array_element in pre_result_rates:
                if (float(element['cat_price']) == float(second_array_element['cat_price'])):
                    first_array.append('duplicate')
                    material_array = []
                    for rate in second_array_element['rates']:
                        if (rate['rate_id'] == element['rates'][0]['rate_id']):
                            material_array.append('duplicate')
                    if(len(material_array) < 1):
                        second_array_element['rates'].append(element['rates'][0])

                else:
                    first_array.append('new')
            control_array = []
            for first_array_element in first_array:
                if(first_array_element == 'duplicate'):
                    control_array.append('duplicate')
            if(len(control_array) < 1):
                pre_result_rates.append(element)
        else:
            pre_result_rates.append(element)


new_struct = []
def array_struct(data):
    global new_struct
    for element in data:
        obj = {}
        obj['cat_price'] = element['cat_price']
        obj['rates'] = []
        obj['rates'].append(element['single_rate'])
        new_struct.append(obj)

    #print(len(pre_result_rates))

null_material_array = []
def null_material(data):
    global null_material_array
    for element in data:
        if(element['rates'][0]['materials'][0]['material_id'] == 'null'):
            element['rates'][0]['materials'] = 'null'
        null_material_array.append(element)

final_array = []
def final_costs(data, aray):
    global final_array
    for element in data:
        f_array = []
        for f_element in aray:
           if(float(element['price']) == float(f_element['cat_price'])):
                f_array.append(f_element['rates'])
        if(len(f_array) > 0):
            element['detail']['rates'] = f_array[0]
        final_array.append(element)

pre_materials = read_obj(costs_list)
pre_rates = rateslist(costs_list)
new_result_rate_materials(pre_result_list_rates)
#данные для отдельной услуги с материалами
#print(len(pre_result_rate_material_list))

array_struct(pre_result_rate_material_list)
null_material(new_struct)

#for i in null_material_array:
 #   print(i)
mulirates(null_material_array)
#for el in pre_result_rates:
  #  print(el)
final_costs(costs_list, null_material_array)
#print(final_array)


#запись в файл работает
final = str(final_array)
final = final.replace("'null'", "null")
final = final.replace("'", '"')
f.write(final)
#print(materials_list)

#with open("new-costs.json", "w", encoding='utf-8') as filenewrates:
    #json.dump(result, filenewrates, sort_keys=False, indent=4, ensure_ascii=False, separators=(',', ': '))
    #filenewrates.close()
# test(costs)

# with open("costs.json", "w", encoding='utf-8') as filecosts:
#   json.dump(costs_list, filecosts, sort_keys=False, indent=4, ensure_ascii=False, separators=(',', ': '))
#  filecosts.close()
# print(costs_list)